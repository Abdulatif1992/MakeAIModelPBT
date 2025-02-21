import joblib
import numpy as np
import os
import pathlib
from openpyxl import load_workbook

# Fayl nomi
script_path = pathlib.Path(__file__).parent.resolve()
fayl_name = script_path / "data" / "1Q_United_C.xlsx"
# Load the model and the scaler from the file
model = joblib.load(script_path / "model" / "1Q_2402_2410.pkl")
scaler = joblib.load(script_path / "model" / "1Q_2402_2410_scaler.pkl")

# Faylni ochish
wb = load_workbook(fayl_name)
ws = wb.active  # Birinchi ishchi varaqni tanlash

# Bashoratlarni yangi ustunga yozish
prediction_column = ws.max_column + 1  # Eng oxirgi ustundan keyingisi
ws.cell(row=4, column=prediction_column, value="Prediction")  # Ustun nomini qo‘shish

for i, row in enumerate(ws.iter_rows(values_only=True)):
    i = i+1
    if(i>=6):
        if row[1] is not None and row[2] is not None and row[4] is not None and row[5] is not None:
            # Sample new data
            s1 = row[1]
            s3 = row[2]
            anchor = row[4]
            new_data = np.array([[s1, s3, anchor]])
            # Scale the new data using the loaded scaler
            scaled_data = scaler.transform(new_data)

            # Make prediction
            prediction = model.predict(scaled_data)
            # print(prediction, f' Prediction: {prediction[0]}')
            # Excel faylga bashorat natijasini yozish
            ws.cell(row=i, column=prediction_column, value=prediction[0])
    

# Yangilangan faylni saqlash
updated_fayl_nomi = script_path / "result" / "1Q_United_C(result).xlsx"
wb.save(updated_fayl_nomi)
print(f"Prediction is saved: {updated_fayl_nomi}")         






