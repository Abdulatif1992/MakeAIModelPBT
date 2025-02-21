import joblib
import numpy as np
import os
import pathlib
from openpyxl import load_workbook

# Fayl nomi
script_path = pathlib.Path(__file__).parent.resolve()
fayl_name = script_path / "data" / "5Q_united 3.xlsx"


# Faylni ochish
wb = load_workbook(fayl_name)
ws = wb.active  # Birinchi ishchi varaqni tanlash



months = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    i = i+1
    if(i>=5):
        if row[1] is not None and row[2] is not None and row[3] is not None and row[5] is not None and row[6] is not None:
            # Sample new data
            testee_month =  str(row[0])[1:5]
            if(testee_month not in months):
               months.append(testee_month)             
            
            # ws.cell(row=i, column=month_column, value=testee_month)
    # if(i>=10):
    #     break        
    

# Bo'sh array yaratamiz
data = []
for month in months:
    count = 0
    sum = 0
    average = 0
    anchor_sum = 0
    anchor_average = 0
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        i = i+1
        if(i>=5):
            if row[1] is not None and row[2] is not None and row[3] is not None and row[5] is not None and row[6] is not None:
                # Sample new data
                testee_month =  str(row[0])[1:5]
                if(testee_month == month):
                    count = count+1  
                    sum = sum + row[4] 
                    anchor_sum = anchor_sum + row[5]        
    
    
    
    average = sum/count
    anchor_average = anchor_sum/count
    data.append([month, count, average, anchor_average])
    


# Bashoratlarni yangi ustunga yozish
month_column = ws.max_column + 1  # Eng oxirgi ustundan keyingisi
ws.cell(row=4, column=month_column, value="month")  # Ustun nomini qo‘shish

anchor_average = ws.max_column + 1  # Eng oxirgi ustundan keyingisi
ws.cell(row=4, column=anchor_average, value="anchorAv")  # Ustun nomini qo‘shish

total_average = ws.max_column + 1  # Eng oxirgi ustundan keyingisi
ws.cell(row=4, column=total_average, value="totalAv")  # Ustun nomini qo‘shish

months = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    i = i+1
    if(i>=5):
        if row[1] is not None and row[2] is not None and row[3] is not None and row[5] is not None and row[6] is not None:
            # Sample new data
            testee_month =  str(row[0])[1:5]
            index = -1 
            for j, row in enumerate(data):  # Har bir qatordan indeks va qiymatni olamiz
                if row[0] == testee_month:  # Agar birinchi ustun qidirilayotgan qiymatga teng bo'lsa
                    index = j  # Indeksni saqlab qo'yamiz
                    break  # Qidiruv tugadi, loopni to'xtatamiz
            ws.cell(row=i, column=total_average, value=data[j][2])    
            ws.cell(row=i, column=anchor_average, value=data[j][3])            
            ws.cell(row=i, column=month_column, value=data[j][0])
    # if(i>=10):
    #     break        
# Yangilangan faylni saqlash
updated_fayl_nomi = script_path / "result" / "5Q_United 3.xlsx"
wb.save(updated_fayl_nomi)
print(f"Prediction is saved: {updated_fayl_nomi}")         






