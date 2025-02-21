import joblib
import numpy as np
import os
import pathlib
from openpyxl import load_workbook

# Fayl nomi
script_path = pathlib.Path(__file__).parent.resolve()
fayl_name = script_path / "data" / "4Q_united_C.xlsx"
# Load the model and the scaler from the file
model = joblib.load(script_path / "model" / "4Q_1602_AncTot.pkl")
scaler = joblib.load(script_path / "model" / "4Q_1602_AncTot_scaler.pkl")

# Faylni ochish
wb = load_workbook(fayl_name)
ws = wb.active  # Birinchi ishchi varaqni tanlash

# Bashoratlarni yangi ustunga yozish
prediction_column = ws.max_column + 1  # Eng oxirgi ustundan keyingisi
ws.cell(row=4, column=prediction_column, value="Prediction")  # Ustun nomini qo‘shish

for i, row in enumerate(ws.iter_rows(values_only=True)):
    i = i+1
    if(i>=5):
        if row[1] is not None and row[2] is not None and row[3] is not None and row[5] is not None and row[6] is not None:
            # Sample new data
            s1s2 = row[1] + row[2]
            s3 = row[3]
            total = row[4]
            anchorAvg = row[8]
            totalAvg = row[9]
            new_data = np.array([[s1s2, s3, total, anchorAvg, totalAvg]])
            # Scale the new data using the loaded scaler
            scaled_data = scaler.transform(new_data)

            # Make prediction
            prediction = model.predict(scaled_data)
            # print(prediction, f' Prediction: {prediction[0]}')
            # Excel faylga bashorat natijasini yozish
            ws.cell(row=i, column=prediction_column, value=prediction[0])
    # if(i>=6610):
    #     break        
    

# Yangilangan faylni saqlash
updated_fayl_nomi = script_path / "result" / "4Q_united_C(1602).xlsx"
wb.save(updated_fayl_nomi)
print(f"Prediction is saved: {updated_fayl_nomi}")         






