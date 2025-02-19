import pathlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report
import joblib

# Fayl nomi
script_path = pathlib.Path(__file__).parent.resolve()

# fayl_name = script_path / "data" / "5Q_United_C 1 Test.xlsx"
fayl_name = script_path / "data" / "4Q_United_C 1.xlsx"

# Faylni ochish
wb = load_workbook(fayl_name)
ws = wb.active  # Birinchi ishchi varaqni tanlash

X = []
Y = []
for i, row in enumerate(ws.iter_rows(values_only=True)):
    i = i+1
    if(i>=5):
        if row[1] is not None and row[2] is not None and row[3] is not None and row[5] is not None and row[6] is not None:
            X.append([row[1] + row[2], row[3], row[5]])
            Y.append(row[6])

    
# Ro‘yxatlarni numpy array yoki pandas DataFrame formatiga o‘tkazish
X = np.array(X)
Y = np.array(Y)

# Agar Y da kategorik qiymatlar bo'lsa, ularni raqamlarga o'zgartirish
if not np.issubdtype(Y.dtype, np.number):
    from sklearn.preprocessing import LabelEncoder
    le = LabelEncoder()
    Y = le.fit_transform(Y)

# Bo'sh qiymatlarni to'ldirish
X = np.nan_to_num(X)  # Barcha NaNlarni 0 bilan almashtirish

# Train-test ajratish
X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.1, random_state=42)


# Standardize the features
scaler = StandardScaler()
X_train = scaler.fit_transform(X_train)
X_test = scaler.transform(X_test)

# Train a Random Forest classifier
model = RandomForestClassifier(n_estimators=100, random_state=42)
model.fit(X_train, y_train)

# Make predictions
y_pred = model.predict(X_test)

# Evaluate the model
accuracy = accuracy_score(y_test, y_pred)
conf_matrix = confusion_matrix(y_test, y_pred)
class_report = classification_report(y_test, y_pred)

print(f'Accuracy: {accuracy}')
print('Confusion Matrix:')
print(conf_matrix)
print('Classification Report:')
print(class_report)

# Save the model and the scaler
joblib.dump(model, script_path / "model" /"4Q_C.pkl")
joblib.dump(scaler, script_path / "model" /"4Q_C_scaler.pkl")